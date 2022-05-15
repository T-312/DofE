import openpyxl
import os

os.chdir("/Users/tim10/Documents")

wb = openpyxl.load_workbook("Movies.xlsx")
ws = wb.active

x = 1
y = 1

while True:
    if ws.cell(x, y).value != None:
        x+=1

    if ws.cell(x, y).value == None:
        print(f"Found empty cell on row {x} in the {y} column!")
        exit()