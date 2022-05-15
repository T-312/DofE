import openpyxl
import os
import random

os.chdir("/Users/tim10/Documents")

wb = openpyxl.load_workbook("Test.xlsx")
ws = wb.active

x = 1
y = 1

while y != 11:
    a = random.randint(0, 1000)
    ws.cell(x, y).value = a
    x+=1
    if x == 101:
        y+=1
        x-=100

    if y == 11:
        wb.save("Test.xlsx")


wb = openpyxl.load_workbook("Test.xlsx")
ws = wb.active

x = 1
y = 1

nums = []

while True:
    a = ws.cell(x, y).value
    nums.append(a)
    x+=1
    if x == 101:
        y+=1
        x-=100

    if y == 11:
        val = 0

        for i in nums:
            val+=int(i)

        print(val)
        exit()