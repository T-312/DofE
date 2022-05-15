import xlsxwriter
import os
import openpyxl
import time
import pandas as pd

username = os.getlogin()
maindir = f"/Users/{username}/Desktop/Text Files/"
os.chdir(maindir)
books = "Books.xlsx"

def titles():
    workbook = xlsxwriter.Workbook("Books.xlsx")
    worksheet = workbook.add_worksheet()
    workbook.close()

    workbook = xlsxwriter.Workbook("Books.xlsx")
    worksheet = workbook.add_worksheet()
    
    print("Creating file...")
    worksheet.write(0, 0, "Time Modified")
    worksheet.write(0, 2, "Book Title")
    worksheet.write(0, 4, "Author name")
    worksheet.write(0, 6, "Publisher")
    worksheet.write(0, 8, "Pages")
    worksheet.write(0, 10, "ISBN")
    workbook.close()

def main():
    wb = openpyxl.load_workbook('Books.xlsx')
    sheet = wb['Sheet1']
    a = 0

    while a != 1:
        rowData = []

        Time = time.asctime()
        rowData.append(Time)
        rowData.append(' ')

        name = input("Enter book name: ")
        rowData.append(name)
        rowData.append(' ')

        author = input("Enter author name: ")
        rowData.append(author)
        rowData.append(' ')

        publisher = input("Enter publisher: ")
        rowData.append(publisher)
        rowData.append(' ')

        pages = int(input("Enter number of pages: "))
        rowData.append(pages)
        rowData.append(' ')

        isbn = input("Enter ISBN: ")
        rowData.append(isbn)
        rowData.append(' ')

        print(rowData)

        sheet.append(rowData)

        cont = input("Continue y/n: ").lower()
        if cont == "y":
            pass

        if cont != "y":
            a+=1
            wb.save('books.xlsx')
            exit()

if books not in os.listdir():
    titles()
    main()

if books in os.listdir():
    main()